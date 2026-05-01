from fastapi import APIRouter, Depends, Query
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime
from database import get_db
from models.product import Product
from models.customer import Customer
from models.supplier import Supplier
from models.inventory import Inventory, InventoryAlert
from models.sales import SalesOrder, SalesStockout, SalesStockoutItem
from models.purchase import PurchaseOrder, PurchaseStockin, PurchaseStockinItem
from models.finance import Receipt, Payment
from schemas.common import ResponseModel
from datetime import datetime, timedelta

router = APIRouter(prefix="/api/reports", tags=["报表"])


# ========== 经营看板 ==========
@router.get("/dashboard", response_model=ResponseModel)
def dashboard(db: Session = Depends(get_db)):
    today = datetime.now().strftime("%Y-%m-%d")
    # 今日销售额
    today_sales = db.query(func.sum(SalesStockout.total_amount)).filter(
        SalesStockout.status == 2,
        SalesStockout.created_at >= today
    ).scalar() or 0
    # 今日采购额
    today_purchase = db.query(func.sum(PurchaseStockin.total_amount)).filter(
        PurchaseStockin.status == 2,
        PurchaseStockin.created_at >= today
    ).scalar() or 0
    # 今日回款
    today_receipt = db.query(func.sum(Receipt.amount)).filter(
        Receipt.status == 1,
        Receipt.created_at >= today
    ).scalar() or 0
    # 今日付款
    today_payment = db.query(func.sum(Payment.amount)).filter(
        Payment.status == 1,
        Payment.created_at >= today
    ).scalar() or 0
    # 库存总量
    invs = db.query(Inventory).all()
    total_stock_qty = sum(i.quantity for i in invs)
    total_stock_value = sum(i.quantity * i.cost_price for i in invs)
    # 应收/应付
    total_receivable = db.query(func.sum(Customer.receivable_balance)).filter(
        Customer.receivable_balance > 0
    ).scalar() or 0
    total_payable = db.query(func.sum(Supplier.payable_balance)).filter(
        Supplier.payable_balance > 0
    ).scalar() or 0
    # 本月订单数
    month_start = datetime.now().strftime("%Y-%m-01")
    month_orders = db.query(SalesOrder).filter(SalesOrder.created_at >= month_start).count()
    # 库存预警数
    alert_count = db.query(InventoryAlert).filter(InventoryAlert.is_handled == 0).count()
    return ResponseModel(data={
        "today_sales": today_sales,
        "today_purchase": today_purchase,
        "today_receipt": today_receipt,
        "today_payment": today_payment,
        "total_stock_qty": total_stock_qty,
        "total_stock_value": total_stock_value,
        "total_receivable": total_receivable,
        "total_payable": total_payable,
        "month_orders": month_orders,
        "alert_count": alert_count,
        "product_count": db.query(Product).count(),
        "customer_count": db.query(Customer).count(),
        "supplier_count": db.query(Supplier).count(),
    })


# ========== 销售报表 ==========
@router.get("/sales", response_model=ResponseModel)
def sales_report(
    group_by: str = Query("day"),  # day/week/month/year
    start_date: str = Query(None), end_date: str = Query(None),
    customer_id: int = Query(None), product_id: int = Query(None),
    db: Session = Depends(get_db)
):
    q = db.query(SalesStockout).filter(SalesStockout.status == 2)
    if start_date:
        q = q.filter(SalesStockout.created_at >= start_date)
    if end_date:
        q = q.filter(SalesStockout.created_at <= datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59))
    if customer_id:
        q = q.filter(SalesStockout.customer_id == customer_id)
    stockouts = q.all()
    # 按日期分组统计
    stats = {}
    for so in stockouts:
        date_key = str(so.created_at)[:10]  # YYYY-MM-DD
        if group_by == "month":
            date_key = date_key[:7]
        elif group_by == "year":
            date_key = date_key[:4]
        if date_key not in stats:
            stats[date_key] = {"date": date_key, "order_count": 0, "total_amount": 0, "cost_amount": 0, "profit": 0}
        stats[date_key]["order_count"] += 1
        stats[date_key]["total_amount"] += so.total_amount
        # 计算成本
        items = db.query(SalesStockoutItem).filter(SalesStockoutItem.stockout_id == so.id).all()
        for item in items:
            if product_id and item.product_id != product_id:
                continue
            inv = db.query(Inventory).filter(Inventory.product_id == item.product_id).first()
            cost = inv.cost_price if inv else 0
            stats[date_key]["cost_amount"] += item.quantity * cost
        stats[date_key]["profit"] = stats[date_key]["total_amount"] - stats[date_key]["cost_amount"]
    result = sorted(stats.values(), key=lambda x: x["date"])
    return ResponseModel(data=result)


# ========== 采购报表 ==========
@router.get("/purchase", response_model=ResponseModel)
def purchase_report(
    group_by: str = Query("day"),
    start_date: str = Query(None), end_date: str = Query(None),
    supplier_id: int = Query(None),
    db: Session = Depends(get_db)
):
    q = db.query(PurchaseStockin).filter(PurchaseStockin.status == 2)
    if start_date:
        q = q.filter(PurchaseStockin.created_at >= start_date)
    if end_date:
        q = q.filter(PurchaseStockin.created_at <= datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59))
    if supplier_id:
        q = q.filter(PurchaseStockin.supplier_id == supplier_id)
    stockins = q.all()
    stats = {}
    for si in stockins:
        date_key = str(si.created_at)[:10]
        if group_by == "month":
            date_key = date_key[:7]
        elif group_by == "year":
            date_key = date_key[:4]
        if date_key not in stats:
            stats[date_key] = {"date": date_key, "order_count": 0, "total_amount": 0}
        stats[date_key]["order_count"] += 1
        stats[date_key]["total_amount"] += si.total_amount
    result = sorted(stats.values(), key=lambda x: x["date"])
    return ResponseModel(data=result)


# ========== 库存报表 ==========
@router.get("/inventory", response_model=ResponseModel)
def inventory_report(warehouse_id: int = Query(None), db: Session = Depends(get_db)):
    q = db.query(Inventory)
    if warehouse_id:
        q = q.filter(Inventory.warehouse_id == warehouse_id)
    invs = q.all()
    result = []
    for inv in invs:
        product = db.query(Product).get(inv.product_id)
        from models.warehouse import Warehouse
        warehouse = db.query(Warehouse).get(inv.warehouse_id)
        result.append({
            "product_id": inv.product_id,
            "product_name": product.name if product else "",
            "product_code": product.code if product else "",
            "warehouse_name": warehouse.name if warehouse else "",
            "quantity": inv.quantity,
            "cost_price": inv.cost_price,
            "total_value": inv.quantity * inv.cost_price
        })
    total_value = sum(r["total_value"] for r in result)
    return ResponseModel(data={"items": result, "total_value": total_value})


# ========== 利润报表 ==========
@router.get("/profit", response_model=ResponseModel)
def profit_report(
    group_by: str = Query("day"),
    start_date: str = Query(None), end_date: str = Query(None),
    db: Session = Depends(get_db)
):
    # 销售收入
    sales_q = db.query(SalesStockout).filter(SalesStockout.status == 2)
    if start_date:
        sales_q = sales_q.filter(SalesStockout.created_at >= start_date)
    if end_date:
        sales_q = sales_q.filter(SalesStockout.created_at <= datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59))
    sales = sales_q.all()
    # 采购成本
    purchase_q = db.query(PurchaseStockin).filter(PurchaseStockin.status == 2)
    if start_date:
        purchase_q = purchase_q.filter(PurchaseStockin.created_at >= start_date)
    if end_date:
        purchase_q = purchase_q.filter(PurchaseStockin.created_at <= datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59))
    purchases = purchase_q.all()
    # 按日期分组
    stats = {}
    for so in sales:
        date_key = str(so.created_at)[:10]
        if group_by == "month":
            date_key = date_key[:7]
        if date_key not in stats:
            stats[date_key] = {"date": date_key, "sales_amount": 0, "cost_amount": 0, "profit": 0, "gross_rate": 0}
        stats[date_key]["sales_amount"] += so.total_amount
        items = db.query(SalesStockoutItem).filter(SalesStockoutItem.stockout_id == so.id).all()
        for item in items:
            inv = db.query(Inventory).filter(Inventory.product_id == item.product_id).first()
            cost = inv.cost_price if inv else 0
            stats[date_key]["cost_amount"] += item.quantity * cost
    for si in purchases:
        date_key = str(si.created_at)[:10]
        if group_by == "month":
            date_key = date_key[:7]
        if date_key not in stats:
            stats[date_key] = {"date": date_key, "sales_amount": 0, "cost_amount": 0, "profit": 0, "gross_rate": 0}
    for key in stats:
        stats[key]["profit"] = stats[key]["sales_amount"] - stats[key]["cost_amount"]
        if stats[key]["sales_amount"] > 0:
            stats[key]["gross_rate"] = round(stats[key]["profit"] / stats[key]["sales_amount"] * 100, 2)
    result = sorted(stats.values(), key=lambda x: x["date"])
    return ResponseModel(data=result)


# ========== Excel导出 ==========
@router.get("/export/sales")
def export_sales(
    start_date: str = Query(None),
    end_date: str = Query(None),
    db: Session = Depends(get_db)
):
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from io import BytesIO

    q = db.query(SalesStockout).filter(SalesStockout.status == 2)
    if start_date:
        q = q.filter(SalesStockout.created_at >= start_date)
    if end_date:
        q = q.filter(SalesStockout.created_at <= datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59))
    orders = q.order_by(SalesStockout.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "销售报表"
    ws.append(["单号", "客户", "仓库", "金额", "日期"])

    from models.customer import Customer as Cust
    from models.warehouse import Warehouse as Wh
    customers_map = {c.id: c.name for c in db.query(Cust).all()}
    warehouses_map = {w.id: w.name for w in db.query(Wh).all()}

    for so in orders:
        ws.append([
            so.code,
            customers_map.get(so.customer_id, ""),
            warehouses_map.get(so.warehouse_id, ""),
            so.total_amount,
            str(so.created_at)[:19]
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sales_report.xlsx"}
    )


@router.get("/export/inventory")
def export_inventory(db: Session = Depends(get_db)):
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from io import BytesIO

    invs = db.query(Inventory).all()
    products_map = {p.id: p for p in db.query(Product).all()}
    from models.warehouse import Warehouse as Wh
    warehouses_map = {w.id: w.name for w in db.query(Wh).all()}

    wb = Workbook()
    ws = wb.active
    ws.title = "库存报表"
    ws.append(["商品编码", "商品名称", "规格", "单位", "仓库", "数量", "成本价", "库存金额"])

    for inv in invs:
        p = products_map.get(inv.product_id)
        ws.append([
            p.code if p else "",
            p.name if p else "",
            p.spec if p else "",
            p.unit if p else "",
            warehouses_map.get(inv.warehouse_id, ""),
            inv.quantity,
            inv.cost_price,
            inv.quantity * inv.cost_price
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inventory_report.xlsx"}
    )


@router.get("/export/finance")
def export_finance(
    start_date: str = Query(None),
    end_date: str = Query(None),
    db: Session = Depends(get_db)
):
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from io import BytesIO

    wb = Workbook()
    # 收款
    ws1 = wb.active
    ws1.title = "收款记录"
    ws1.append(["单号", "客户", "金额", "日期"])

    rq = db.query(Receipt).filter(Receipt.status == 1)
    if start_date:
        rq = rq.filter(Receipt.created_at >= start_date)
    if end_date:
        rq = rq.filter(Receipt.created_at <= datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59))

    customers_map = {c.id: c.name for c in db.query(Customer).all()}
    for r in rq.all():
        ws1.append([r.code, customers_map.get(r.customer_id, ""), r.amount, str(r.created_at)[:19]])

    # 付款
    ws2 = wb.create_sheet("付款记录")
    ws2.append(["单号", "供应商", "金额", "日期"])

    pq = db.query(Payment).filter(Payment.status == 1)
    if start_date:
        pq = pq.filter(Payment.created_at >= start_date)
    if end_date:
        pq = pq.filter(Payment.created_at <= datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59))

    suppliers_map = {s.id: s.name for s in db.query(Supplier).all()}
    for p in pq.all():
        ws2.append([p.code, suppliers_map.get(p.supplier_id, ""), p.amount, str(p.created_at)[:19]])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=finance_report.xlsx"}
    )


# ========== 趋势图 ==========
@router.get("/trend", response_model=ResponseModel)
def trend_report(
    trend_type: str = Query("sales"),  # sales/purchase
    period: str = Query("month"),  # month/quarter
    months: int = Query(12, ge=1, le=24),
    db: Session = Depends(get_db)
):
    """采购/销售趋势数据（按月/季度）"""
    from collections import OrderedDict
    now = datetime.now()
    data = OrderedDict()

    if trend_type == "sales":
        # 查询已确认的销售出库单
        records = db.query(SalesStockout).filter(SalesStockout.status == 2).all()
        for r in records:
            dt = r.created_at
            if not dt:
                continue
            if period == "month":
                key = dt.strftime("%Y-%m")
            elif period == "quarter":
                q = (dt.month - 1) // 3 + 1
                key = f"{dt.year}-Q{q}"
            else:
                key = dt.strftime("%Y-%m")
            if key not in data:
                data[key] = {"amount": 0, "count": 0}
            data[key]["amount"] += r.total_amount or 0
            data[key]["count"] += 1
    else:
        records = db.query(PurchaseStockin).filter(PurchaseStockin.status == 2).all()
        for r in records:
            dt = r.created_at
            if not dt:
                continue
            if period == "month":
                key = dt.strftime("%Y-%m")
            elif period == "quarter":
                q = (dt.month - 1) // 3 + 1
                key = f"{dt.year}-Q{q}"
            else:
                key = dt.strftime("%Y-%m")
            if key not in data:
                data[key] = {"amount": 0, "count": 0}
            data[key]["amount"] += r.total_amount or 0
            data[key]["count"] += 1

    # 填充空月份/季度
    result = []
    if period == "month":
        for i in range(months - 1, -1, -1):
            dt = now - timedelta(days=i * 30)
            key = dt.strftime("%Y-%m")
            d = data.get(key, {"amount": 0, "count": 0})
            result.append({"period": key, "amount": round(d["amount"], 2), "count": d["count"]})
    else:
        for i in range(months // 3):
            dt = now - timedelta(days=i * 90)
            q = (dt.month - 1) // 3 + 1
            key = f"{dt.year}-Q{q}"
            if key not in [r["period"] for r in result]:
                d = data.get(key, {"amount": 0, "count": 0})
                result.insert(0, {"period": key, "amount": round(d["amount"], 2), "count": d["count"]})

    return ResponseModel(data=result)
