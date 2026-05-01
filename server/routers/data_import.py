import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from sqlalchemy.orm import Session
from database import get_db
from models.product import Product
from models.customer import Customer
from models.supplier import Supplier
from models.inventory import Inventory
from models.warehouse import Warehouse
from schemas.common import ResponseModel
from openpyxl import load_workbook

router = APIRouter(prefix="/api/data-import", tags=["数据导入"])

IMPORT_TYPES = {
    "products": {"label": "商品", "fields": ["编码", "条码", "名称", "规格", "单位", "进货价", "零售价", "库存下限", "库存上限"]},
    "customers": {"label": "客户", "fields": ["编码", "名称", "联系人", "电话", "地址", "等级", "信用额度", "期初应收"]},
    "suppliers": {"label": "供应商", "fields": ["编码", "名称", "联系人", "电话", "地址", "账期", "期初应付"]},
    "inventory": {"label": "库存", "fields": ["仓库编码", "商品编码", "数量", "成本价"]},
}


def parse_excel(file_bytes: bytes):
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    rows = []
    headers = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip() if c else "" for c in row]
        else:
            rows.append(dict(zip(headers, row)))
    wb.close()
    return headers, rows


@router.get("/types", response_model=ResponseModel)
def get_import_types():
    data = [{"value": k, "label": v["label"], "fields": v["fields"]} for k, v in IMPORT_TYPES.items()]
    return ResponseModel(data=data)


@router.get("/template/{import_type}")
def download_template(import_type: str):
    if import_type not in IMPORT_TYPES:
        raise HTTPException(status_code=400, detail="无效导入类型")
    from fastapi.responses import StreamingResponse
    wb = __import__("openpyxl").Workbook()
    ws = wb.active
    ws.title = IMPORT_TYPES[import_type]["label"] + "导入模板"
    fields = IMPORT_TYPES[import_type]["fields"]
    ws.append(fields)
    # Add example row
    examples = {
        "products": ["SP001", "6901234567890", "统一冰红茶500ml", "500ml*24瓶", "箱", 48, 60, 10, 500],
        "customers": ["KH001", "好邻居便利店", "张三", "13800138000", "武汉市武昌区", "A", 5000, 1560],
        "suppliers": ["GYS001", "统一企业", "李四", "13900139000", "上海市", "月结30天", 0],
        "inventory": ["WH001", "SP001", 100, 48],
    }
    ws.append(examples.get(import_type, [""] * len(fields)))
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    from urllib.parse import quote
    filename = IMPORT_TYPES[import_type]["label"] + "导入模板.xlsx"
    encoded = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"}
    )


@router.post("/preview/{import_type}", response_model=ResponseModel)
async def preview_import(import_type: str, file: UploadFile = File(...), db: Session = Depends(get_db)):
    if import_type not in IMPORT_TYPES:
        raise HTTPException(status_code=400, detail="无效导入类型")
    content = await file.read()
    try:
        headers, rows = parse_excel(content)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel解析失败: {str(e)}")

    if not rows:
        raise HTTPException(status_code=400, detail="Excel中无数据行")

    # Validate and build preview
    errors = []
    valid = []
    for i, row in enumerate(rows):
        row_num = i + 2  # Excel row (1-indexed, header is row 1)
        result = validate_row(import_type, row, db)
        if result["errors"]:
            errors.append({"row": row_num, "errors": result["errors"]})
        else:
            valid.append({"row": row_num, "data": result["data"]})

    return ResponseModel(data={
        "headers": headers,
        "total_rows": len(rows),
        "valid_count": len(valid),
        "error_count": len(errors),
        "valid": valid[:100],  # preview first 100
        "errors": errors[:50],
    })


@router.post("/execute/{import_type}", response_model=ResponseModel)
async def execute_import(import_type: str, file: UploadFile = File(...), db: Session = Depends(get_db)):
    if import_type not in IMPORT_TYPES:
        raise HTTPException(status_code=400, detail="无效导入类型")
    content = await file.read()
    try:
        headers, rows = parse_excel(content)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel解析失败: {str(e)}")

    errors = []
    imported = 0
    for i, row in enumerate(rows):
        row_num = i + 2
        result = validate_row(import_type, row, db)
        if result["errors"]:
            errors.append({"row": row_num, "errors": result["errors"]})
            continue
        try:
            do_import(import_type, result["data"], db)
            imported += 1
        except Exception as e:
            errors.append({"row": row_num, "errors": [str(e)]})

    db.commit()
    return ResponseModel(data={
        "total_rows": len(rows),
        "imported": imported,
        "errors": errors,
    })


def validate_row(import_type: str, row: dict, db: Session) -> dict:
    errors = []
    data = {}

    if import_type == "products":
        code = str(row.get("编码", "")).strip()
        name = str(row.get("名称", "")).strip()
        if not code:
            errors.append("编码不能为空")
        if not name:
            errors.append("名称不能为空")
        if code and db.query(Product).filter(Product.code == code).first():
            errors.append(f"编码{code}已存在")
        data = {
            "code": code,
            "barcode": str(row.get("条码", "") or "").strip(),
            "name": name,
            "spec": str(row.get("规格", "") or "").strip(),
            "unit": str(row.get("单位", "") or "").strip(),
            "purchase_price": float(row.get("进货价", 0) or 0),
            "retail_price": float(row.get("零售价", 0) or 0),
            "stock_min": float(row.get("库存下限", 0) or 0),
            "stock_max": float(row.get("库存上限", 0) or 0),
        }

    elif import_type == "customers":
        code = str(row.get("编码", "")).strip()
        name = str(row.get("名称", "")).strip()
        if not code:
            errors.append("编码不能为空")
        if not name:
            errors.append("名称不能为空")
        if code and db.query(Customer).filter(Customer.code == code).first():
            errors.append(f"编码{code}已存在")
        data = {
            "code": code,
            "name": name,
            "contact": str(row.get("联系人", "") or "").strip(),
            "phone": str(row.get("电话", "") or "").strip(),
            "address": str(row.get("地址", "") or "").strip(),
            "level": str(row.get("等级", "") or "").strip(),
            "credit_limit": float(row.get("信用额度", 0) or 0),
            "receivable_balance": float(row.get("期初应收", 0) or 0),
        }

    elif import_type == "suppliers":
        code = str(row.get("编码", "")).strip()
        name = str(row.get("名称", "")).strip()
        if not code:
            errors.append("编码不能为空")
        if not name:
            errors.append("名称不能为空")
        if code and db.query(Supplier).filter(Supplier.code == code).first():
            errors.append(f"编码{code}已存在")
        data = {
            "code": code,
            "name": name,
            "contact": str(row.get("联系人", "") or "").strip(),
            "phone": str(row.get("电话", "") or "").strip(),
            "address": str(row.get("地址", "") or "").strip(),
            "payment_term": str(row.get("账期", "") or "").strip(),
            "payable_balance": float(row.get("期初应付", 0) or 0),
        }

    elif import_type == "inventory":
        wh_code = str(row.get("仓库编码", "")).strip()
        prod_code = str(row.get("商品编码", "")).strip()
        if not wh_code:
            errors.append("仓库编码不能为空")
        if not prod_code:
            errors.append("商品编码不能为空")
        warehouse = db.query(Warehouse).filter(Warehouse.code == wh_code).first() if wh_code else None
        product = db.query(Product).filter(Product.code == prod_code).first() if prod_code else None
        if wh_code and not warehouse:
            errors.append(f"仓库编码{wh_code}不存在")
        if prod_code and not product:
            errors.append(f"商品编码{prod_code}不存在")
        data = {
            "warehouse_id": warehouse.id if warehouse else None,
            "product_id": product.id if product else None,
            "quantity": float(row.get("数量", 0) or 0),
            "cost_price": float(row.get("成本价", 0) or 0),
        }

    return {"errors": errors, "data": data}


def do_import(import_type: str, data: dict, db: Session):
    if import_type == "products":
        db.add(Product(**data))
    elif import_type == "customers":
        db.add(Customer(**data))
    elif import_type == "suppliers":
        db.add(Supplier(**data))
    elif import_type == "inventory":
        inv = db.query(Inventory).filter(
            Inventory.warehouse_id == data["warehouse_id"],
            Inventory.product_id == data["product_id"]
        ).first()
        if inv:
            inv.quantity = data["quantity"]
            inv.cost_price = data["cost_price"]
        else:
            db.add(Inventory(**data))
